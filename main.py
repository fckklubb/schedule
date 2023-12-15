from sys import argv
from datetime import datetime, date, time
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from configs import sheet_prefix, months, m_days, YY, days, five_day_return, shift_day_return, SCHEDULE_FILE


if __name__ == '__main__':
    txts = {'input': 'Plz, enter month (1-12 or 0 for all months) OR exit, if the work is done: ',
            'error': 'Nope! Can\'t accept it :( Try to type \'exit\' if the work is done\nOR enter a correct month number, specifically 1-12 or 0 for all months.. Let\'s try again..',
            'y_n': 'Are U suuuuure?? Type \'y\' or \'n\', plz: ',
            'y_n_only': 'Nope! Type \'y\' or \'n\' only!..',
            'dontwannado': 'Ufff, nice.. Don\'t wanna do it..',
            'start': 'OK, fine.. Starting to set all the year schedule..'}
    while True:
        option = input(txts['input'])
        option = option.lower()
        if option == 'exit': break
        try:
            option = int(option)
        except ValueError: 
            print(txts['error'])
            continue
        if option == 0:
            
            while True:
                answer = input(txts['y_n'])
                answer = answer.lower()
                if answer == 'n':
                    print(txts['dontwannado'])
                    break    
                elif answer == 'y':
                    print(txts['start'])
                    wb = load_workbook(filename = SCHEDULE_FILE)

                    for m in range(1, 13):
                        print(f'Working on the month nr. {m}..')
                        ws = wb[f"{m}"]
                        ws['A13'] = sheet_prefix + f" {months[m-1]} {YY}"
                        for d in range(1, m_days[m-1]+1):
                            day = date(year=YY, month=m, day=d)
                            ws.cell(row=14, column=d+4).value = d
                            ws.cell(row=13, column=d+4).value = days[date(YY, m, d).isoweekday()-1]

                            # print('Month: ', m, 'day: ', d)

                            #Shift0
                            ws.cell(row=15, column=d+4).value, ws.cell(row=16, column=d+4).value, ws.cell(row=17, column=d+4).value = five_day_return(day)
                            #Shift1
                            x, y, z = shift_day_return(day, 1)
                            ws.cell(row=19, column=d+4).value, ws.cell(row=26, column=d+4).value, ws.cell(row=27, column=d+4).value = x, y, z
                            ws.cell(row=20, column=d+4).value, ws.cell(row=29, column=d+4).value, ws.cell(row=30, column=d+4).value = x, y, z
                            ws.cell(row=47, column=d+4).value, ws.cell(row=52, column=d+4).value, ws.cell(row=53, column=d+4).value = x, y, z
                            #Shift2
                            x, y, z = shift_day_return(day, 2)
                            ws.cell(row=21, column=d+4).value, ws.cell(row=32, column=d+4).value, ws.cell(row=33, column=d+4).value = x, y, z
                            ws.cell(row=22, column=d+4).value, ws.cell(row=35, column=d+4).value, ws.cell(row=36, column=d+4).value = x, y, z
                            ws.cell(row=48, column=d+4).value, ws.cell(row=55, column=d+4).value, ws.cell(row=56, column=d+4).value = x, y, z

            
                        wb.save(filename = SCHEDULE_FILE)
                        wb.close()
                        print(f'The work on the month nr. {m} is done!')
                    break
                else:
                    print(txts['y_n_only'])

        elif option < 13:
            wb = load_workbook(filename = SCHEDULE_FILE)
            m = option
            print(f'Working on the month nr. {m}..')
            ws = wb[f"{m}"]
            ws['A13'] = sheet_prefix + f" {months[m-1]} {YY}"
            for d in range(1, m_days[m-1]+1):
                day = date(year=YY, month=m, day=d)
                ws.cell(row=14, column=d+4).value = d
                ws.cell(row=13, column=d+4).value = days[date(YY, m, d).isoweekday()-1]

                print('Month: ', m, 'day: ', d)

                #Shift0
                ws.cell(row=15, column=d+4).value, ws.cell(row=16, column=d+4).value, ws.cell(row=17, column=d+4).value = five_day_return(day)
                #Shift1
                x, y, z = shift_day_return(day, 1)
                ws.cell(row=19, column=d+4).value, ws.cell(row=26, column=d+4).value, ws.cell(row=27, column=d+4).value = x, y, z
                ws.cell(row=20, column=d+4).value, ws.cell(row=29, column=d+4).value, ws.cell(row=30, column=d+4).value = x, y, z
                ws.cell(row=47, column=d+4).value, ws.cell(row=52, column=d+4).value, ws.cell(row=53, column=d+4).value = x, y, z
                #Shift2
                x, y, z = shift_day_return(day, 2)
                ws.cell(row=21, column=d+4).value, ws.cell(row=32, column=d+4).value, ws.cell(row=33, column=d+4).value = x, y, z
                ws.cell(row=22, column=d+4).value, ws.cell(row=35, column=d+4).value, ws.cell(row=36, column=d+4).value = x, y, z
                ws.cell(row=48, column=d+4).value, ws.cell(row=55, column=d+4).value, ws.cell(row=56, column=d+4).value = x, y, z


            wb.save(filename = SCHEDULE_FILE)
            wb.close()
        else:
            print(txts['error'])